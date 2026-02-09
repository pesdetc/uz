"""Модуль для экспорта результатов в Excel"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os
import config


class ExcelExporter:
    """Экспорт результатов в Excel с форматированием"""
    
    def __init__(self):
        self.output_dir = config.OUTPUT_DIR
        self.filename = config.OUTPUT_FILENAME
    
    def prepare_data(self, usernames_data: list, whois_results: list) -> pd.DataFrame:
        """
        Подготовка данных для экспорта
        
        Args:
            usernames_data: Данные о юзернеймах
            whois_results: Результаты WHOIS-проверки
            
        Returns:
            DataFrame с объединенными данными
        """
        # Создаем DataFrame из юзернеймов
        df_users = pd.DataFrame(usernames_data)
        
        # Создаем DataFrame из WHOIS результатов
        df_whois = pd.DataFrame(whois_results)
        
        # Объединяем данные
        # Создаем ключ для объединения
        df_users['domain'] = df_users['username'].str.lower() + '.uz'
        
        # Объединяем по домену
        df_merged = pd.merge(
            df_users,
            df_whois,
            on='domain',
            how='left'
        )
        
        # Отбираем нужные колонки и переименовываем
        df_final = df_merged[[
            'source',
            'username',
            'url',
            'domain',
            'status',
            'expiry_date',
            'created_date',
            'registrar'
        ]].copy()
        
        # Переименовываем колонки на русский
        df_final.columns = [
            'Источник',
            'Username',
            'URL профиля',
            'Домен .uz',
            'Статус',
            'Дата истечения',
            'Дата регистрации',
            'Регистратор'
        ]
        
        # Заменяем значения статуса на русский
        status_map = {
            'Available': '✅ Свободен',
            'Registered': '❌ Занят',
            'Unknown': '❓ Неизвестно',
            'Error': '⚠️ Ошибка'
        }
        
        df_final['Статус'] = df_final['Статус'].map(status_map).fillna(df_final['Статус'])
        
        return df_final
    
    def apply_formatting(self, filepath: str):
        """
        Применение форматирования к Excel-файлу
        
        Args:
            filepath: Путь к файлу Excel
        """
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Стили
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        
        available_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        registered_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Форматируем заголовки
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Форматируем данные
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                
                # Подсвечиваем статус
                if cell.column == 5:  # Колонка "Статус"
                    if '✅ Свободен' in str(cell.value):
                        cell.fill = available_fill
                    elif '❌ Занят' in str(cell.value):
                        cell.fill = registered_fill
        
        # Автоширина колонок
        column_widths = {
            'A': 12,  # Источник
            'B': 20,  # Username
            'C': 40,  # URL профиля
            'D': 20,  # Домен
            'E': 15,  # Статус
            'F': 15,  # Дата истечения
            'G': 15,  # Дата регистрации
            'H': 25   # Регистратор
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Устанавливаем высоту строк
        ws.row_dimensions[1].height = 30
        
        # Закрепляем первую строку
        ws.freeze_panes = 'A2'
        
        wb.save(filepath)
    
    def export(self, usernames_data: list, whois_results: list) -> str:
        """
        Экспорт данных в Excel
        
        Args:
            usernames_data: Данные о юзернеймах
            whois_results: Результаты WHOIS-проверки
            
        Returns:
            Путь к созданному файлу
        """
        # Создаем директорию если не существует
        os.makedirs(self.output_dir, exist_ok=True)
        
        # Генерируем имя файла с датой
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"uz_domains_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        print(f"\n📊 Создание отчета...")
        
        # Подготавливаем данные
        df = self.prepare_data(usernames_data, whois_results)
        
        # Экспортируем в Excel
        df.to_excel(filepath, index=False, engine='openpyxl')
        
        # Применяем форматирование
        self.apply_formatting(filepath)
        
        print(f"✅ Отчет сохранен: {filepath}")
        print(f"📈 Всего записей: {len(df)}")
        print(f"   ✅ Свободных доменов: {len(df[df['Статус'] == '✅ Свободен'])}")
        print(f"   ❌ Занятых доменов: {len(df[df['Статус'] == '❌ Занят'])}")
        
        return filepath
